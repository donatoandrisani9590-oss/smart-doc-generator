"""
Bulk Generation API: CSV/Excel import and batch document generation

Features:
- CSV and Excel (.xlsx) file support
- Pre-import validation with detailed error reporting
- Field-level validation based on FormField definitions
- Progress tracking and job management
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update
from pydantic import BaseModel
from typing import Optional, List, Any
import csv
import io
import os
import re
import tempfile
from datetime import datetime

from app.db import get_db
from app.models.enterprise import BulkJob, FormField
from app.models.documents import DocumentType
from app.api.deps import get_current_user
from app.tasks.bulk_tasks import process_bulk_job

router = APIRouter()


import logging

logger = logging.getLogger(__name__)

# Maximum file size for upload (10 MB)
MAX_UPLOAD_SIZE = 10 * 1024 * 1024


def cleanup_temp_file(path: str) -> None:
    """Background task to delete temporary file."""
    try:
        if os.path.exists(path):
            os.remove(path)
            logger.debug(f"Cleaned up temp file: {path}")
    except OSError as e:
        logger.warning(f"Failed to cleanup temp file {path}: {e}")


class FieldError(BaseModel):
    row: int
    column: str
    value: str
    error: str


class BulkValidateResponse(BaseModel):
    is_valid: bool
    row_count: int
    column_count: int
    errors: List[str]
    warnings: List[str]
    field_errors: List[FieldError]
    preview_data: List[dict]  # First 5 rows for preview
    detected_columns: List[str]
    missing_required_columns: List[str]
    unknown_columns: List[str]


@router.get("/templates/{document_type_id}")
async def download_template(
    document_type_id: int,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    background_tasks: BackgroundTasks = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Download a template with the correct columns for a document type.

    Supports CSV and Excel formats.
    Includes:
    - All form fields as columns
    - Example data row
    - Comments for validation rules
    """
    # Get document type
    doc_type = await db.get(DocumentType, document_type_id)
    if not doc_type:
        raise HTTPException(status_code=404, detail="Dokumenttyp nicht gefunden")

    # Get form fields
    result = await db.execute(
        select(FormField)
        .where(FormField.document_type_id == document_type_id)
        .order_by(FormField.display_order.nulls_last())
    )
    fields = result.scalars().all()

    if not fields:
        # Fallback to basic fields if no form fields defined
        fields_data = [
            {"name": "vorname", "label": "Vorname", "required": True, "example": "Max"},
            {"name": "nachname", "label": "Nachname", "required": True, "example": "Mustermann"},
            {"name": "eintrittsdatum", "label": "Eintrittsdatum", "required": False, "example": "01.01.2025"},
        ]
    else:
        fields_data = []
        for f in fields:
            example = f.default_value or ""
            if not example:
                if f.field_type == "date":
                    example = "01.01.2025"
                elif f.field_type == "number":
                    example = "0"
                elif f.field_type == "email":
                    example = "beispiel@firma.de"
                elif f.field_type == "checkbox":
                    example = "ja"
                else:
                    example = ""

            fields_data.append({
                "name": f.field_name,
                "label": f.field_label,
                "required": f.is_required,
                "example": example,
                "type": f.field_type,
                "help": f.help_text or "",
            })

    if format == "xlsx":
        # Excel format
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daten"

            # Header row
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, field in enumerate(fields_data, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = field["name"]
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            # Example row
            for col, field in enumerate(fields_data, 1):
                ws.cell(row=2, column=col).value = field["example"]

            # Instructions sheet
            ws_help = wb.create_sheet("Anleitung")
            ws_help.cell(row=1, column=1).value = "Spalte"
            ws_help.cell(row=1, column=2).value = "Beschreibung"
            ws_help.cell(row=1, column=3).value = "Pflicht"
            ws_help.cell(row=1, column=4).value = "Typ"
            ws_help.cell(row=1, column=5).value = "Hinweis"

            for row, field in enumerate(fields_data, 2):
                ws_help.cell(row=row, column=1).value = field["name"]
                ws_help.cell(row=row, column=2).value = field["label"]
                ws_help.cell(row=row, column=3).value = "Ja" if field["required"] else "Nein"
                ws_help.cell(row=row, column=4).value = field.get("type", "text")
                ws_help.cell(row=row, column=5).value = field.get("help", "")

            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Tempfile mit Auto-Cleanup verwenden
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                f.write(buffer.getvalue())
                temp_path = f.name

            # Cleanup nach Response
            if background_tasks:
                background_tasks.add_task(cleanup_temp_file, temp_path)

            return FileResponse(
                temp_path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"vorlage_{doc_type.name.replace(' ', '_')}.xlsx"
            )

        except ImportError:
            raise HTTPException(
                status_code=400,
                detail="Excel-Export nicht verfügbar. Bitte CSV verwenden."
            )

    else:
        # CSV format
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')

        # Header
        writer.writerow([f["name"] for f in fields_data])

        # Example row
        writer.writerow([f["example"] for f in fields_data])

        # Tempfile mit Auto-Cleanup verwenden
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8-sig") as f:
            f.write(output.getvalue())
            temp_path = f.name

        # Cleanup nach Response
        if background_tasks:
            background_tasks.add_task(cleanup_temp_file, temp_path)

        return FileResponse(
            temp_path,
            media_type="text/csv",
            filename=f"vorlage_{doc_type.name.replace(' ', '_')}.csv"
        )


def parse_file_content(content: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """Parse CSV or Excel file content into rows."""
    rows = []
    columns = []

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Excel file
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            if data:
                columns = [str(c).strip().lower() if c else f"spalte_{i}" for i, c in enumerate(data[0])]
                for row_data in data[1:]:
                    row = {}
                    for i, val in enumerate(row_data):
                        if i < len(columns):
                            row[columns[i]] = str(val) if val is not None else ""
                    rows.append(row)
        except ImportError:
            raise HTTPException(status_code=400, detail="Excel-Unterstützung nicht verfügbar. Bitte CSV verwenden.")
    else:
        # CSV file
        try:
            decoded = content.decode('utf-8')
        except UnicodeDecodeError:
            decoded = content.decode('latin-1')

        reader = csv.DictReader(io.StringIO(decoded), delimiter=';')  # German CSV often uses semicolon
        rows = list(reader)
        if not rows:
            # Try with comma delimiter
            reader = csv.DictReader(io.StringIO(decoded), delimiter=',')
            rows = list(reader)

        if rows:
            columns = [k.lower().strip() for k in rows[0].keys()]
            # Normalize keys
            rows = [{k.lower().strip(): v for k, v in row.items()} for row in rows]

    return rows, columns


def validate_field_value(value: str, field: FormField) -> tuple[bool, str]:
    """Validate a single field value against FormField rules."""
    if not value and field.is_required:
        return False, "Pflichtfeld ist leer"

    if not value:
        return True, ""

    # Type validation
    if field.field_type == "number":
        try:
            num = float(value.replace(',', '.'))
            if field.min_value is not None and num < field.min_value:
                return False, f"Wert muss mindestens {field.min_value} sein"
            if field.max_value is not None and num > field.max_value:
                return False, f"Wert darf maximal {field.max_value} sein"
        except ValueError:
            return False, "Ungültiger Zahlenwert"

    elif field.field_type == "date":
        # Try common date formats
        date_formats = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"]
        valid_date = False
        for fmt in date_formats:
            try:
                datetime.strptime(value, fmt)
                valid_date = True
                break
            except ValueError:
                continue
        if not valid_date:
            return False, "Ungültiges Datumsformat (erwartet: TT.MM.JJJJ oder JJJJ-MM-TT)"

    elif field.field_type == "email":
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, value):
            return False, "Ungültige E-Mail-Adresse"

    elif field.field_type == "checkbox":
        valid_values = ["true", "false", "ja", "nein", "yes", "no", "1", "0", "x", ""]
        if value.lower() not in valid_values:
            return False, "Ungültiger Wert für Checkbox (erwartet: ja/nein, true/false, 1/0)"

    # Text length validation
    if field.field_type in ["text", "textarea"]:
        if field.min_length and len(value) < field.min_length:
            return False, f"Text muss mindestens {field.min_length} Zeichen haben"
        if field.max_length and len(value) > field.max_length:
            return False, f"Text darf maximal {field.max_length} Zeichen haben"

    # Pattern validation
    if field.pattern:
        try:
            if not re.match(field.pattern, value):
                return False, field.pattern_error_message or "Format stimmt nicht mit dem erwarteten Muster überein"
        except re.error:
            pass  # Invalid pattern, skip validation

    # Select options validation
    if field.field_type == "select" and field.options:
        try:
            import json
            options = json.loads(field.options)
            valid_values = [str(opt.get("value", opt)).lower() for opt in options] if isinstance(options, list) else []
            if value.lower() not in valid_values:
                return False, f"Ungültige Option. Erlaubt: {', '.join(valid_values)}"
        except (json.JSONDecodeError, TypeError):
            pass

    return True, ""


@router.post("/validate")
async def validate_upload(
    document_type_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
) -> BulkValidateResponse:
    """
    Validate a CSV or Excel file before bulk generation.

    Performs:
    - File format detection (CSV/Excel)
    - Column mapping to form fields
    - Row-by-row field validation
    - Required field checks
    - Type and format validation

    Returns detailed errors and warnings with row/column references.
    """
    # Validate file size first (avoid reading huge files into memory)
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"Datei zu groß. Maximum: {MAX_UPLOAD_SIZE // (1024*1024)} MB"
        )

    filename = file.filename or "upload.csv"

    # Validate file extension
    if not filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Ungültiges Dateiformat. Erlaubt: CSV, XLSX, XLS"
        )

    # Parse file
    try:
        rows, detected_columns = parse_file_content(content, filename)
    except Exception as e:
        return BulkValidateResponse(
            is_valid=False,
            row_count=0,
            column_count=0,
            errors=[f"Datei konnte nicht gelesen werden: {str(e)}"],
            warnings=[],
            field_errors=[],
            preview_data=[],
            detected_columns=[],
            missing_required_columns=[],
            unknown_columns=[],
        )

    errors = []
    warnings = []
    field_errors = []

    if not rows:
        errors.append("Datei ist leer oder enthält keine Datenzeilen")
        return BulkValidateResponse(
            is_valid=False,
            row_count=0,
            column_count=len(detected_columns),
            errors=errors,
            warnings=warnings,
            field_errors=[],
            preview_data=[],
            detected_columns=detected_columns,
            missing_required_columns=[],
            unknown_columns=[],
        )

    # Get form fields for document type
    result = await db.execute(
        select(FormField)
        .where(FormField.document_type_id == document_type_id)
        .order_by(FormField.display_order.nulls_last())
    )
    form_fields = result.scalars().all()

    # Build field map (field_name -> FormField)
    field_map = {f.field_name.lower(): f for f in form_fields}
    required_fields = [f.field_name.lower() for f in form_fields if f.is_required]

    # Check for missing required columns
    missing_required = [f for f in required_fields if f not in detected_columns]

    # Check for unknown columns
    known_fields = set(field_map.keys())
    unknown_columns = [c for c in detected_columns if c not in known_fields]

    if missing_required:
        errors.append(f"Fehlende Pflichtfelder: {', '.join(missing_required)}")

    if unknown_columns:
        warnings.append(f"Unbekannte Spalten (werden ignoriert): {', '.join(unknown_columns)}")

    # Validate each row
    for row_idx, row in enumerate(rows, start=2):  # Start at 2 (1 = header)
        for col_name, value in row.items():
            if col_name in field_map:
                field = field_map[col_name]
                is_valid, error_msg = validate_field_value(value, field)
                if not is_valid:
                    field_errors.append(FieldError(
                        row=row_idx,
                        column=col_name,
                        value=value[:50] if value else "",
                        error=error_msg,
                    ))

        # Limit errors per file
        if len(field_errors) >= 100:
            warnings.append("Validierung abgebrochen nach 100 Fehlern. Bitte Datei korrigieren.")
            break

    # Build preview (first 5 rows)
    preview_data = rows[:5]

    # Determine if valid
    is_valid = len(errors) == 0 and len(field_errors) == 0

    return BulkValidateResponse(
        is_valid=is_valid,
        row_count=len(rows),
        column_count=len(detected_columns),
        errors=errors,
        warnings=warnings,
        field_errors=field_errors[:50],  # Limit displayed errors
        preview_data=preview_data,
        detected_columns=detected_columns,
        missing_required_columns=missing_required,
        unknown_columns=unknown_columns,
    )


@router.post("/generate")
async def start_bulk_generation(
    document_type_id: int,
    signatory_name: str = Query(..., min_length=2, description="Pflichtfeld: Name des Unterzeichners (DIN 5008)"),
    file: UploadFile = File(...),
    output_format: str = "pdf",
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Start a bulk generation job with DIN 5008 format.
    Returns job ID for progress polling.
    """
    content = await file.read()

    # Validate file size
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"Datei zu groß. Maximum: {MAX_UPLOAD_SIZE // (1024*1024)} MB"
        )

    # Validate output format
    if output_format not in ["pdf", "docx"]:
        raise HTTPException(
            status_code=400,
            detail="Ungültiges Ausgabeformat. Erlaubt: pdf, docx"
        )

    try:
        decoded = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            decoded = content.decode('latin-1')
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=400,
                detail="Datei konnte nicht dekodiert werden. Bitte UTF-8 Encoding verwenden."
            )

    # Try semicolon first (German CSV), then comma
    reader = csv.DictReader(io.StringIO(decoded), delimiter=';')
    rows = list(reader)
    if not rows or not rows[0]:
        reader = csv.DictReader(io.StringIO(decoded), delimiter=',')
        rows = list(reader)

    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Datei ist leer oder enthält keine gültigen Daten"
        )
    
    # Create job record
    job = BulkJob(
        status="PENDING",
        total_records=len(rows),
        processed_records=0,
        created_by_id=current_user.id
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)
    
    # Queue Celery task with DIN 5008 signatory name
    process_bulk_job.delay(
        job_id=job.id,
        document_type_id=document_type_id,
        rows=[dict(r) for r in rows],
        output_format=output_format,
        signatory_name=signatory_name
    )
    
    return {
        "job_id": job.id,
        "status": "queued",
        "total_records": len(rows)
    }


@router.get("/jobs")
async def list_jobs(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """List all bulk jobs for the current user."""
    result = await db.execute(
        select(BulkJob)
        .where(BulkJob.created_by_id == current_user.id)
        .order_by(BulkJob.created_at.desc())
        .limit(50)
    )
    jobs = result.scalars().all()
    
    return [
        {
            "id": j.id,
            "status": j.status,
            "total_records": j.total_records,
            "processed_records": j.processed_records,
            "progress_percent": round((j.processed_records / j.total_records) * 100) if j.total_records > 0 else 0,
            "created_at": j.created_at.isoformat() if j.created_at else None,
        }
        for j in jobs
    ]


@router.get("/jobs/{job_id}")
async def get_job_status(
    job_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get detailed status of a bulk job (for progress polling)."""
    job = await db.get(BulkJob, job_id)
    
    if not job or job.created_by_id != current_user.id:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return {
        "id": job.id,
        "status": job.status,
        "total_records": job.total_records,
        "processed_records": job.processed_records,
        "progress_percent": round((job.processed_records / job.total_records) * 100) if job.total_records > 0 else 0,
        "result_file_path": job.result_file_path,
        "created_at": job.created_at.isoformat() if job.created_at else None,
    }


@router.get("/jobs/{job_id}/download")
async def download_job_result(
    job_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Download the resulting ZIP file from a completed bulk job."""
    job = await db.get(BulkJob, job_id)
    
    if not job or job.created_by_id != current_user.id:
        raise HTTPException(status_code=404, detail="Job not found")
    
    if job.status != "COMPLETED":
        raise HTTPException(status_code=400, detail=f"Job not complete. Status: {job.status}")
    
    if not job.result_file_path or not os.path.exists(job.result_file_path):
        raise HTTPException(status_code=404, detail="Result file not found")
    
    return FileResponse(
        job.result_file_path,
        media_type="application/zip",
        filename=f"bulk_dokumente_{job_id}.zip"
    )


@router.delete("/jobs/{job_id}")
async def cancel_job(
    job_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Cancel a running bulk job.

    Implementiert gemäß v4.2 Spezifikation Kapitel 15.4:
    - Setzt den Job-Status auf CANCELLED
    - Der Celery-Task prüft regelmäßig diesen Status und bricht ab
    - Bereits generierte Dateien werden aufgeräumt

    Der Abbruch erfolgt "graceful" - der Task beendet die aktuelle
    Dokumentgenerierung und stoppt dann.

    WICHTIG: Verwendet atomic update um Race Conditions zu vermeiden.
    """
    # Atomic Update mit Bedingung um Race Conditions zu vermeiden
    # Nur Jobs die noch PENDING oder PROCESSING sind können abgebrochen werden
    result = await db.execute(
        update(BulkJob)
        .where(
            BulkJob.id == job_id,
            BulkJob.created_by_id == current_user.id,
            BulkJob.status.in_(["PENDING", "PROCESSING"])
        )
        .values(status="CANCELLED")
        .returning(BulkJob.id, BulkJob.processed_records, BulkJob.total_records)
    )

    updated_row = result.fetchone()

    if not updated_row:
        # Job nicht gefunden oder bereits beendet - prüfen warum
        job = await db.get(BulkJob, job_id)

        if not job or job.created_by_id != current_user.id:
            raise HTTPException(status_code=404, detail="Job not found")

        # Job existiert aber hat falschen Status
        raise HTTPException(
            status_code=400,
            detail=f"Job bereits beendet (Status: {job.status})"
        )

    await db.commit()

    return {
        "status": "cancelled",
        "message": "Abbruch wurde angefordert. Der Job wird nach Abschluss der aktuellen Verarbeitung gestoppt.",
        "processed_records": updated_row.processed_records,
        "total_records": updated_row.total_records
    }
